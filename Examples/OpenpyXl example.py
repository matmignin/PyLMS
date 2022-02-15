
from openpyxl import Workbook, load_workbook
import warnings
warnings.simplefilter("ignore")

wb = load_workbook('Formulations_Workbook.xlsm')
ws = wb.active


from openpyxl.utils import get_column_letter

for row in range(8,22):
	for col in range(2,3): # get 2nd Colomn
		char = get_column_letter(col) # to get the column letter name
		print(ws[char + str(row)].value) # print out the cell value, then the
